"""
Session management endpoints (teacher-facing):
  - POST /sessions              (create a new session)
  - GET  /sessions              (list sessions I own / am enrolled in)
  - POST /sessions/{id}/enroll  (enroll a student by email)
  - GET  /sessions/{id}/attendance  (attendance register for a session)
  - GET  /sessions/{id}/attendance/export  (download Excel)
  - GET  /sessions/{id}/analytics   (session stats)
"""

import secrets as _secrets
import io
from datetime import datetime, timezone
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from pydantic import BaseModel

from app.db.session import get_db
from app.models.session import Session
from app.models.user import User, UserRole
from app.models.enrollment import Enrollment
from app.models.attendance import Attendance
from app.api.deps import get_current_user, require_teacher
from app.core.roll_no import normalize_roll_no


class CreateSessionRequest(BaseModel):
    name: str
    start_time: datetime
    end_time: datetime
    latitude: float | None = None
    longitude: float | None = None
    radius_meters: int = 100


class EnrollRequest(BaseModel):
    student_email: str

class BulkEnrollRequest(BaseModel):
    student_emails: list[str]

class JoinSessionRequest(BaseModel):
    code: str


class StudentUpdateRequest(BaseModel):
    roll_no: str | None = None
    email: str | None = None
    device_id: str | None = None

router = APIRouter(prefix="/sessions", tags=["Sessions"])


def _pick_best_attendance(records: list[Attendance]) -> dict[UUID, Attendance]:
    record_map: dict[UUID, Attendance] = {}
    for record in records:
        if record.session_id not in record_map or record.status == "Present":
            record_map[record.session_id] = record
    return record_map


def _attendance_priority(status: str) -> int:
    if status == "Present":
        return 3
    if status == "Leave":
        return 2
    if status == "Out of Bounds":
        return 1
    return 0


def _status_for_matrix(status: str) -> str:
    if status == "Present":
        return "Present"
    if status == "Leave":
        return "Leave"
    return "Absent"


@router.post("")
async def create_session(
    payload: CreateSessionRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_teacher),
):
    session = Session(
        name=payload.name,
        start_time=payload.start_time,
        end_time=payload.end_time,
        secret_key=_secrets.token_hex(32),
        instructor_id=current_user.id,
        latitude=payload.latitude,
        longitude=payload.longitude,
        radius_meters=payload.radius_meters,
    )
    db.add(session)
    await db.commit()
    await db.refresh(session)

    return {
        "id": str(session.id),
        "name": session.name,
        "enrollment_code": session.enrollment_code,
        "start_time": session.start_time.isoformat(),
        "end_time": session.end_time.isoformat(),
        "latitude": session.latitude,
        "longitude": session.longitude,
        "radius_meters": session.radius_meters,
    }


@router.get("")
async def list_my_sessions(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role == UserRole.TEACHER:
        query = select(Session).where(Session.instructor_id == current_user.id).order_by(Session.start_time.desc())
    else:
        query = (
            select(Session)
            .join(Enrollment, Enrollment.session_id == Session.id)
            .where(Enrollment.student_id == current_user.id)
            .order_by(Session.start_time.desc())
        )

    result = await db.execute(query)
    sessions = result.scalars().all()

    return [
        {
            "id": str(s.id),
            "name": s.name,
            "enrollment_code": s.enrollment_code,
            "start_time": s.start_time.isoformat(),
            "end_time": s.end_time.isoformat(),
            "latitude": s.latitude,
            "longitude": s.longitude,
            "radius_meters": s.radius_meters,
        }
        for s in sessions
    ]


@router.get("/student/stats")
async def get_student_stats(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Returns the student's attendance stats grouped by session name.
    """
    if current_user.role == UserRole.TEACHER:
        raise HTTPException(status_code=403, detail="Only students can view these stats.")

    session_query = (
        select(Session)
        .join(Enrollment, Enrollment.session_id == Session.id)
        .where(Enrollment.student_id == current_user.id)
        .order_by(Session.start_time.desc())
    )
    session_result = await db.execute(session_query)
    sessions = session_result.scalars().all()

    session_ids = [session.id for session in sessions]
    attendance_map: dict[UUID, Attendance] = {}
    if session_ids:
        attendance_query = select(Attendance).where(
            Attendance.student_id == current_user.id,
            Attendance.session_id.in_(session_ids),
        )
        attendance_result = await db.execute(attendance_query)
        attendance_records = attendance_result.scalars().all()
        attendance_map = _pick_best_attendance(attendance_records)

    stats: dict[str, dict[str, int]] = {}
    total_classes: int = 0
    total_attended: int = 0
    attendance_register = []

    for session in sessions:
        attendance = attendance_map.get(session.id)
        status = attendance.status if attendance else "Absent"
        name = session.name
        if name not in stats:
            stats[name] = {"total_classes": 0, "attended": 0, "missed": 0}

        stats[name]["total_classes"] += 1
        total_classes += 1

        if status == "Present":
            stats[name]["attended"] += 1
            total_attended += 1
        else:
            stats[name]["missed"] += 1

        attendance_register.append({
            "session_id": str(session.id),
            "session_name": session.name,
            "date": session.start_time.isoformat(),
            "status": status,
            "scan_time": attendance.timestamp.isoformat() if attendance else None,
            "roll_no": current_user.roll_no or "N/A",
            "email": current_user.email,
            "device_id": current_user.device_id or "NOT BOUND",
        })

    return {
        "profile": {
            "student_id": str(current_user.id),
            "full_name": current_user.full_name,
            "email": current_user.email,
            "roll_no": current_user.roll_no or "N/A",
            "device_id": current_user.device_id or "NOT BOUND",
        },
        "overall": {
            "total_classes": total_classes,
            "total_attended": total_attended,
            "total_missed": total_classes - total_attended,
        },
        "by_course": [{"course_name": k, **v} for k, v in stats.items()],
        "attendance_register": attendance_register,
    }


@router.post("/join")
async def join_session(
    payload: JoinSessionRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role != UserRole.STUDENT:
        raise HTTPException(status_code=403, detail="Only students can join sessions via code.")

    code_upper = payload.code.strip().upper()
    result = await db.execute(select(Session).where(Session.enrollment_code == code_upper))
    session = result.scalar_one_or_none()
    
    if not session:
        raise HTTPException(status_code=404, detail="Invalid enrollment code. Session not found.")

    result = await db.execute(
        select(Enrollment).where(
            Enrollment.student_id == current_user.id,
            Enrollment.session_id == session.id,
        )
    )
    if result.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="You are already enrolled in this session.")

    enrollment = Enrollment(student_id=current_user.id, session_id=session.id)
    db.add(enrollment)
    await db.commit()

    return {"message": f"Successfully joined {session.name}!", "session_id": str(session.id)}


@router.post("/{session_id}/enroll")
async def enroll_student(
    session_id: UUID,
    payload: EnrollRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_teacher),
):
    result = await db.execute(select(Session).where(Session.id == session_id))
    session = result.scalar_one_or_none()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    if session.instructor_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not your session")

    result = await db.execute(select(User).where(User.email == payload.student_email))
    student = result.scalar_one_or_none()
    if not student:
        raise HTTPException(status_code=404, detail=f"No user found with email {payload.student_email}")

    result = await db.execute(
        select(Enrollment).where(
            Enrollment.student_id == student.id,
            Enrollment.session_id == session_id,
        )
    )
    if result.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Student already enrolled")

    enrollment = Enrollment(student_id=student.id, session_id=session_id)
    db.add(enrollment)
    await db.commit()

    return {"message": f"{student.full_name} enrolled successfully", "student_email": student.email}


@router.post("/{session_id}/enroll/bulk")
async def bulk_enroll_students(
    session_id: UUID,
    payload: BulkEnrollRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_teacher),
):
    result = await db.execute(select(Session).where(Session.id == session_id))
    session = result.scalar_one_or_none()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    if session.instructor_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not your session")

    # Get all users by emails
    result = await db.execute(select(User).where(User.email.in_(payload.student_emails)))
    students = result.scalars().all()
    
    if not students:
        raise HTTPException(status_code=404, detail="No users found with the provided emails")

    student_ids = [student.id for student in students]
    result = await db.execute(
        select(Enrollment).where(
            Enrollment.student_id.in_(student_ids),
            Enrollment.session_id == session_id,
        )
    )
    existing_enrollments = {e.student_id for e in result.scalars().all()}

    new_enrollments = []
    enrolled_emails = []
    for student in students:
        if student.id not in existing_enrollments:
            new_enrollments.append(Enrollment(student_id=student.id, session_id=session_id))
            enrolled_emails.append(student.email)

    if new_enrollments:
        db.add_all(new_enrollments)
        await db.commit()

    return {
        "message": f"Successfully enrolled {len(new_enrollments)} new students.",
        "enrolled_emails": enrolled_emails,
        "not_found_count": len(payload.student_emails) - len(students),
        "already_enrolled_count": len(students) - len(new_enrollments)
    }


# ---------------------------------------------------------------------------
# GET /sessions/{session_id}/attendance/matrix — Date-wise class register
# ---------------------------------------------------------------------------
@router.get("/{session_id}/attendance/matrix")
async def get_session_attendance_matrix(
    session_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_teacher),
):
    # Verify ownership of selected class/session
    selected_session_q = select(Session).where(Session.id == session_id)
    selected_session_res = await db.execute(selected_session_q)
    selected_session = selected_session_res.scalar_one_or_none()
    if not selected_session:
        raise HTTPException(status_code=404, detail="Session not found")
    if selected_session.instructor_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not your session")

    # A class register spans all teacher sessions with the same class name.
    class_sessions_q = (
        select(Session)
        .where(
            Session.instructor_id == current_user.id,
            Session.name == selected_session.name,
        )
        .order_by(Session.start_time.asc())
    )
    class_sessions_result = await db.execute(class_sessions_q)
    class_sessions = class_sessions_result.scalars().all()
    class_session_ids = [sess.id for sess in class_sessions]

    # Students are scoped to the selected class/session only.
    students_q = (
        select(User.id, User.roll_no, User.full_name, User.email, User.device_id)
        .join(Enrollment, Enrollment.student_id == User.id)
        .where(Enrollment.session_id == session_id)
    )
    students_result = await db.execute(students_q)
    students = students_result.all()

    if not class_sessions:
        return {
            "session_id": str(session_id),
            "session_name": selected_session.name,
            "columns": [],
            "rows": [],
            "summary": {"total_students": 0, "total_dates": 0},
        }

    if not students:
        return {
            "session_id": str(session_id),
            "session_name": selected_session.name,
            "columns": [
                {
                    "session_id": str(sess.id),
                    "date": sess.start_time.isoformat(),
                }
                for sess in class_sessions
            ],
            "rows": [],
            "summary": {"total_students": 0, "total_dates": len(class_sessions)},
        }

    student_ids = [student_id for student_id, *_ in students]

    enrollments_q = select(Enrollment.student_id, Enrollment.session_id).where(
        Enrollment.student_id.in_(student_ids),
        Enrollment.session_id.in_(class_session_ids),
    )
    enrollments_result = await db.execute(enrollments_q)
    enrollments = enrollments_result.all()
    enrollment_pairs = {(student_id, class_session_id) for student_id, class_session_id in enrollments}

    attendance_q = select(Attendance).where(
        Attendance.student_id.in_(student_ids),
        Attendance.session_id.in_(class_session_ids),
    )
    attendance_result = await db.execute(attendance_q)
    attendance_records = attendance_result.scalars().all()

    attendance_map: dict[tuple[UUID, UUID], Attendance] = {}
    for record in attendance_records:
        key = (record.student_id, record.session_id)
        existing = attendance_map.get(key)
        if not existing or _attendance_priority(record.status) > _attendance_priority(existing.status):
            attendance_map[key] = record

    columns = [
        {
            "session_id": str(sess.id),
            "date": sess.start_time.isoformat(),
        }
        for sess in class_sessions
    ]

    rows = []
    for student_id, roll_no, full_name, email, device_id in students:
        total_classes = 0
        total_present = 0

        records = []
        for sess in class_sessions:
            is_enrolled = (student_id, sess.id) in enrollment_pairs
            if not is_enrolled:
                records.append({
                    "session_id": str(sess.id),
                    "status": None,
                })
                continue

            total_classes += 1
            attendance = attendance_map.get((student_id, sess.id))
            status = _status_for_matrix(attendance.status if attendance else "Absent")
            if status == "Present":
                total_present += 1
            records.append({
                "session_id": str(sess.id),
                "status": status,
            })

        rows.append({
            "student_id": str(student_id),
            "roll_no": roll_no or "N/A",
            "full_name": full_name,
            "email": email,
            "device_id": device_id or "NOT BOUND",
            "total_classes": total_classes,
            "total_attended": total_present,
            "total_missed": total_classes - total_present,
            "attendance_rate": round(total_present / total_classes * 100) if total_classes > 0 else 0,
            "records": records,
        })

    rows.sort(key=lambda row: (row["roll_no"] == "N/A", row["roll_no"], row["full_name"]))

    return {
        "session_id": str(session_id),
        "session_name": selected_session.name,
        "columns": columns,
        "rows": rows,
        "summary": {
            "total_students": len(rows),
            "total_dates": len(columns),
        },
    }


# ---------------------------------------------------------------------------
# GET /sessions/{id}/attendance — Attendance register (like a class register)
# ---------------------------------------------------------------------------
@router.get("/{session_id}/attendance")
async def get_session_attendance(
    session_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_teacher),
):
    # Verify session ownership
    result = await db.execute(select(Session).where(Session.id == session_id))
    session = result.scalar_one_or_none()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    if session.instructor_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not your session")

    # Get all enrolled students
    enrolled_q = (
        select(User.id, User.full_name, User.email, User.device_id, User.roll_no)
        .join(Enrollment, Enrollment.student_id == User.id)
        .where(Enrollment.session_id == session_id)
    )
    enrolled_result = await db.execute(enrolled_q)
    enrolled_students = enrolled_result.all()

    # Get attendance records for this session
    att_q = select(Attendance).where(Attendance.session_id == session_id)
    att_result = await db.execute(att_q)
    attendance_records = att_result.scalars().all()

    # Build lookup: student_id -> attendance record
    att_map = {}
    for a in attendance_records:
        if a.student_id not in att_map or a.status == "Present":
            att_map[a.student_id] = a

    # Build register rows
    register = []
    for student_id, full_name, email, device_id, roll_no in enrolled_students:
        att = att_map.get(student_id)
        register.append({
            "student_id": str(student_id),
            "roll_no": roll_no or "N/A",
            "full_name": full_name,
            "email": email,
            "device_id": device_id or "NOT BOUND",
            "status": att.status if att else "Absent",
            "scan_time": att.timestamp.isoformat() if att else None,
            "scan_lat": att.scan_latitude if att else None,
            "scan_lon": att.scan_longitude if att else None,
        })

    return {
        "session_id": str(session_id),
        "session_name": session.name,
        "total_enrolled": len(enrolled_students),
        "total_present": sum(1 for r in register if r["status"] == "Present"),
        "total_absent": sum(1 for r in register if r["status"] == "Absent"),
        "register": register,
    }


# ---------------------------------------------------------------------------
# GET /sessions/{id}/attendance/export — Download as Excel
# ---------------------------------------------------------------------------
@router.get("/{session_id}/attendance/export")
async def export_attendance_excel(
    session_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_teacher),
):
    # Reuse the attendance logic
    result = await db.execute(select(Session).where(Session.id == session_id))
    session = result.scalar_one_or_none()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    if session.instructor_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not your session")

    # Get enrolled students
    enrolled_q = (
        select(User.id, User.full_name, User.email, User.device_id)
        .join(Enrollment, Enrollment.student_id == User.id)
        .where(Enrollment.session_id == session_id)
    )
    enrolled_result = await db.execute(enrolled_q)
    enrolled_students = enrolled_result.all()

    # Get attendance
    att_q = select(Attendance).where(Attendance.session_id == session_id)
    att_result = await db.execute(att_q)
    attendance_records = att_result.scalars().all()

    att_map = {}
    for a in attendance_records:
        if a.student_id not in att_map or a.status == "Present":
            att_map[a.student_id] = a

    # Build Excel
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Register"

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = f"Attendance Register — {session.name}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Session ID: {session.id} | Date: {session.start_time.strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Headers
    headers = ["#", "Full Name", "Email", "Device ID", "Status", "Scan Time", "Location"]
    header_fill = PatternFill(start_color="37352F", end_color="37352F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Data rows
    present_fill = PatternFill(start_color="EDF3EC", end_color="EDF3EC", fill_type="solid")
    absent_fill = PatternFill(start_color="FBE4E4", end_color="FBE4E4", fill_type="solid")

    for idx, (student_id, full_name, email, device_id) in enumerate(enrolled_students, 1):
        att = att_map.get(student_id)
        row = idx + 4
        status = att.status if att else "Absent"
        fill = present_fill if status == "Present" else absent_fill

        values = [
            idx,
            full_name,
            email,
            device_id or "NOT BOUND",
            status,
            att.timestamp.strftime("%H:%M:%S") if att else "—",
            f"{att.scan_latitude:.4f}, {att.scan_longitude:.4f}" if att and att.scan_latitude else "—",
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    # Summary row
    summary_row = len(enrolled_students) + 6
    ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=summary_row, column=4, value=f"Enrolled: {len(enrolled_students)}").font = Font(bold=True)
    present_count = sum(1 for sid, *_ in enrolled_students if sid in att_map and att_map[sid].status == "Present")
    ws.cell(row=summary_row, column=5, value=f"Present: {present_count}").font = Font(bold=True, color="0F7B6C")

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 22

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"attendance_{session.name.replace(' ', '_')}_{session.start_time.strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# GET /sessions/students/all — All students the teacher teaches
# ---------------------------------------------------------------------------
@router.get("/students/all")
async def get_all_my_students(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_teacher),
):
    teacher_sessions_q = select(Session).where(Session.instructor_id == current_user.id).order_by(Session.start_time.desc())
    teacher_sessions_result = await db.execute(teacher_sessions_q)
    teacher_sessions = teacher_sessions_result.scalars().all()
    teacher_session_ids = [s.id for s in teacher_sessions]

    if not teacher_session_ids:
        return []

    # Get all students enrolled in any of the above sessions
    students_q = (
        select(User)
        .join(Enrollment, Enrollment.student_id == User.id)
        .where(Enrollment.session_id.in_(teacher_session_ids))
        .distinct()
    )
    students_result = await db.execute(students_q)
    students = students_result.scalars().all()

    # Get attendance for all these sessions
    att_q = select(Attendance).where(Attendance.session_id.in_(teacher_session_ids))
    att_result = await db.execute(att_q)
    attendances = att_result.scalars().all()

    enrollment_q = select(Enrollment).where(Enrollment.session_id.in_(teacher_session_ids))
    enrollment_res = await db.execute(enrollment_q)
    enrollments = enrollment_res.scalars().all()

    stats = []
    for s in students:
        s_enrollments = [e for e in enrollments if e.student_id == s.id]
        total_classes = len(s_enrollments)
        s_att = [a for a in attendances if a.student_id == s.id and a.status == "Present"]
        total_attended = len(s_att)
        
        # Build chronological history
        history = []
        for sess in teacher_sessions:
            # Only include if they were enrolled
            if any(e.session_id == sess.id for e in s_enrollments):
                a_record = next((a for a in attendances if a.student_id == s.id and a.session_id == sess.id), None)
                status = a_record.status if a_record else "Absent"
                history.append({
                    "session_id": str(sess.id),
                    "session_name": sess.name,
                    "date": sess.start_time.isoformat(),
                    "status": status,
                })

        stats.append({
            "student_id": str(s.id),
            "roll_no": s.roll_no or "N/A",
            "full_name": s.full_name,
            "email": s.email,
            "device_id": s.device_id or "NOT BOUND",
            "total_classes": total_classes,
            "total_attended": total_attended,
            "total_missed": total_classes - total_attended,
            "attendance_rate": round(total_attended / total_classes * 100) if total_classes > 0 else 0,
            "history": history
        })

    return sorted(stats, key=lambda x: x["full_name"])


# ---------------------------------------------------------------------------
# PUT /sessions/students/{student_id} — Teacher updates student data
# ---------------------------------------------------------------------------
@router.put("/students/{student_id}")
async def update_student(
    student_id: UUID,
    payload: StudentUpdateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_teacher),
):
    # Verify the teacher actually teaches this student
    teacher_sessions_q = select(Session.id).where(Session.instructor_id == current_user.id)
    teacher_sessions_res = await db.execute(teacher_sessions_q)
    teacher_sessions = [row[0] for row in teacher_sessions_res.all()]

    if not teacher_sessions:
        raise HTTPException(status_code=403, detail="You do not teach any sessions.")

    ver_q = select(Enrollment).where(
        Enrollment.student_id == student_id,
        Enrollment.session_id.in_(teacher_sessions)
    )
    ver_res = await db.execute(ver_q)
    if not ver_res.scalars().first():
        raise HTTPException(status_code=403, detail="Student is not in any of your sessions.")

    # Get user
    user_q = select(User).where(User.id == student_id)
    user_res = await db.execute(user_q)
    student = user_res.scalar_one_or_none()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found.")

    if payload.roll_no is not None:
        normalized_roll_no = normalize_roll_no(payload.roll_no, required=False)
        if normalized_roll_no:
            roll_no_owner_q = select(User.id).where(
                User.roll_no == normalized_roll_no,
                User.id != student_id,
            )
            roll_no_owner_res = await db.execute(roll_no_owner_q)
            if roll_no_owner_res.scalar_one_or_none():
                raise HTTPException(status_code=400, detail="Roll number already registered.")
        student.roll_no = normalized_roll_no
    if payload.email is not None:
        cleaned_email = payload.email.strip().lower()
        if not cleaned_email:
            raise HTTPException(status_code=400, detail="Email cannot be empty.")
        email_owner_q = select(User.id).where(
            User.email == cleaned_email,
            User.id != student_id,
        )
        email_owner_res = await db.execute(email_owner_q)
        if email_owner_res.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="Email already registered.")
        student.email = cleaned_email
    if payload.device_id is not None:
        cleaned_device_id = payload.device_id.strip()
        if cleaned_device_id.lower() in ["", "null", "none", "not bound"]:
            student.device_id = None
        else:
            student.device_id = cleaned_device_id

    try:
        await db.commit()
    except IntegrityError as exc:
        await db.rollback()
        err_text = str(exc.orig).lower() if getattr(exc, "orig", None) else str(exc).lower()
        if "roll_no" in err_text:
            raise HTTPException(status_code=400, detail="Roll number already registered.")
        if "email" in err_text:
            raise HTTPException(status_code=400, detail="Email already registered.")
        raise HTTPException(status_code=400, detail="Student details could not be updated.")

    return {
        "message": "Student details updated successfully.",
        "student": {
            "id": str(student.id),
            "roll_no": student.roll_no,
            "email": student.email,
            "device_id": student.device_id,
        },
    }
